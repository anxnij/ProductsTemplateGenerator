import os
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import date
from copy import copy
 
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
 
SOURCE_SHEET = "TPR new"
TEMPLATE_SHEET = "For check"
STYLE_SOURCE_SHEET = "Discounts"
TARGET_GREEN_RGB = "FF00FF99"
 
CODES = [
    "SC", "R4", "R2", "SA", "S1",
    "SB", "R0", "R1", "S9",
    "15578177", "15363754"
]
 
OPERATIONS = [
    "Discount creation",
    "Modification",
    "Deletion"
]
 
CUSTOMER_NAMES = {
    "SC": "Barnaul",
    "R4": "Ekaterinburg",
    "R2": "Kemerovo",
    "SA": "Krasnoyarsk",
    "S1": "Moscow",
    "SB": "Novokuznetsk",
    "R0": "Novosibirsk",
    "R1": "Omsk",
    "S9": "Tula",
    "15578177": "MSK KIOSK. ROZNISA GT-IC",
    "15363754": "MSK OSTROV MECHTY GT-IC"
}
DATA_FILL_RGB="FFD9D9D9"
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
DATA_FILL= PatternFill(fill_type="solid", start_color=DATA_FILL_RGB, end_color=DATA_FILL_RGB)
 
def next_month_first_day():
    today = date.today()
    if today.month == 12:
        return date(today.year + 1, 1, 1)
    return date(today.year, today.month + 1, 1)
 
def normalize_percent(value):
    if value is None:
        return 0.0
 
    if isinstance(value, str):
        value = value.replace("%", "").replace(",", ".").strip()
        if value == "":
            return 0.0
        try:
            value = float(value)
        except Exception:
            return 0.0
 
    try:
        value = float(value)
    except Exception:
        return 0.0
 
    if -1 < value < 1 and value != 0:
        value *= 100
 
    return abs(value)
 
def normalize_code(code):
    if code is None:
        return ""
    return str(code).strip().upper().replace(" ", "")
 
def build_output_path_from_template(template_path, g6_text):
    template_dir = os.path.dirname(template_path)
    file_name = f"Темплейт новинки_TPR GT_{g6_text}.xlsx"
    return os.path.join(template_dir, file_name)
 
def create_template_copy(template_path, output_path):
    if os.path.abspath(template_path) == os.path.abspath(output_path):
        raise ValueError("Итоговый файл совпадает с шаблоном. Нельзя перезаписать шаблон.")
    shutil.copy2(template_path, output_path)
 
def is_target_green(cell):
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return False
 
    color = fill.fgColor
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        return False
 
    return str(rgb).upper() == TARGET_GREEN_RGB
 
def collect_green_rows(ws_src):
    rows = []
 
    for r in range(5, ws_src.max_row + 1):
        material = ws_src.cell(r, 6).value  # F
        if material is None or str(material).strip() == "":
            continue
 
        if is_target_green(ws_src.cell(r, 6)):
            rows.append(r)
 
    return rows
 
def build_customer_map(ws):
    """
    Карта customer code -> колонка
    по строке 3 в диапазоне Q:X.
    Поддерживает обычные и составные коды вроде R2,SB.
    """
    result = {}
 
    for col in range(17, 25):  # Q:X
        raw = ws.cell(3, col).value
        if raw is None:
            continue
 
        raw_norm = normalize_code(raw)
        if not raw_norm:
            continue
 
        result[raw_norm] = col
 
        parts = [p for p in raw_norm.split(",") if p]
        if len(parts) > 1:
            for part in parts:
                result[part] = col
 
            if set(parts) == {"R2", "SB"}:
                result["R2,SB"] = col
                result["SB,R2"] = col
 
    return result
 
def get_y603_value(ws_src, row_idx, code, customer_map):
    """
    Особые правила:
    15578177 -> S1
    15363754 -> S1
    """
    lookup_code = normalize_code(code)
 
    if lookup_code in ("15578177", "15363754"):
        lookup_code = "S1"
 
    col = customer_map.get(lookup_code)
    if col is None:
        return 0.0
 
    value = ws_src.cell(row_idx, col).value
    return normalize_percent(value)
 
def build_nomenclature_map(ws_src):
    """
    В шаблон в E пишется код из F источника.
    Номенклатура лежит в H источника.
    Значит для J нужен словарь F -> H.
    """
    result = {}
 
    for r in range(5, ws_src.max_row + 1):
        material_code = ws_src.cell(r, 6).value   
        nomenclature = ws_src.cell(r, 8).value    
 
        if material_code is None:
            continue
 
        key = str(material_code).strip()
        if key and key not in result:
            result[key] = nomenclature
 
    return result

def fill_nomenclature_formula(ws_for_check, path: str):
    """
    VLOOKUP(E9, ''!$F:$H, 3, 0)
    """
    print(path)   
    
    formatted_path = path.replace(os.path.basename(path), "") + f"[{os.path.basename(path)}]" + "TPR new"
    print(formatted_path)  
    
    for i in range(9, ws_for_check.max_row + 1):
        if ws_for_check[f"E{i}"].value is not None:
            ws_for_check[f"J{i}"] = f"=VLOOKUP(E{i}, '{formatted_path}'!$F:$H, 3, 0)"

 
def clear_old_rows(ws):
    if ws.max_row > 8:
        ws.delete_rows(9, ws.max_row - 8)
 
def fill_static_part(ws, operation, g6_text):
    ws["B3"] = "irina.yarlykova@unirusgroup.ru"
    ws["E3"] = date.today().strftime("%d.%m.%Y")
    ws["H3"] = operation
 
    ws["A6"] = "R001-UL RUSSIA RETAIL"
    ws["B6"] = "ZURU"
    cell = ws["C6"]
    cell.value = 40002341
    cell.number_format = "0"
    ws["E6"] = "On-invoice_Russia"
    ws["G6"] = g6_text
    ws["H6"] = "31.12.9999"
 
def apply_data_row_style(wb_out, ws_out, row_idx):
    """
    Копируем стиль строки данных с листа Discounts.
    Так заливка будет такой же, как на Discounts.
    """
    ws_style = wb_out[STYLE_SOURCE_SHEET]
    style_row = 9  
 
    for col in range(1, 11):  # A:J
        src = ws_style.cell(style_row, col)
        dst = ws_out.cell(row_idx, col)
 
        if src.font:
            dst.font = copy(src.font)
 
        if src.border:
            dst.border = copy(src.border)
 
        if src.number_format:
            dst.number_format = src.number_format
 
        if src.protection:
            dst.protection = copy(src.protection)
 
        dst.alignment = copy(CENTER_ALIGNMENT)
 
    if ws_style.row_dimensions[style_row].height:
        ws_out.row_dimensions[row_idx].height = ws_style.row_dimensions[style_row].height
 
def force_data_fill(ws, start_row=9, end_row=None):
    if end_row is None:
        end_row = ws.max_row
 
    for row_idx in range(start_row, end_row + 1):
        for col in range(1, 9):  
            cell = ws.cell(row_idx, col)
            cell.fill = copy(DATA_FILL)
            cell.alignment = copy(CENTER_ALIGNMENT)
            
def write_code_row(ws, wb_out, row_idx, condition_type, code, material, amount, nomenclature_map):
    apply_data_row_style(wb_out, ws, row_idx)
 
    ws.cell(row_idx, 1).value = condition_type
 
    ws.cell(row_idx, 2).value = (
        "Customer number" if str(code).isdigit()
        else "Customer group"
    )
 
    code_cell = ws.cell(row_idx, 3)
    if str(code).isdigit():
        code_cell.value = int(code)
        code_cell.number_format = "0"
    else:
        code_cell.value = code
 
    ws.cell(row_idx, 4).value = ""
 
    material_cell = ws.cell(row_idx, 5)
    try:
        material_cell.value = int(str(material).strip())
        material_cell.number_format = "0"
    except Exception:
        material_cell.value = material
 
    amount_cell = ws.cell(row_idx, 7)
    amount_cell.value = round(float(amount or 0), 2)
    amount_cell.number_format = "0.00"
 
    ws.cell(row_idx, 8).value = "%"
 
    ws.cell(row_idx, 9).value = CUSTOMER_NAMES.get(str(code), "")

    material_key = "" if material is None else str(material).strip()
    ws.cell(row_idx, 10).value = nomenclature_map.get(material_key, "")
 
    for col in range(1, 11):
        ws.cell(row_idx, col).alignment = copy(CENTER_ALIGNMENT)
 
def write_block(ws, wb_out, start_row, material, base_percent, y603_map, nomenclature_map):
 
    for i, code in enumerate(CODES):
        row_idx = start_row + i
        write_code_row(
            ws=ws,
            wb_out=wb_out,
            row_idx=row_idx,
            condition_type="Y602",
            code=code,
            material=material,
            amount=base_percent,
            nomenclature_map=nomenclature_map
        )
 
    for i, code in enumerate(CODES):
        row_idx = start_row + len(CODES) + i
        write_code_row(
            ws=ws,
            wb_out=wb_out,
            row_idx=row_idx,
            condition_type="Y603",
            code=code,
            material=material,
            amount=y603_map.get(code, 0),
            nomenclature_map=nomenclature_map
        )

 
def delete_zero_rows(ws):
    for row_idx in range(ws.max_row, 8, -1):
        value = ws.cell(row_idx, 7).value
        try:
            if value is None or float(value) == 0:
                ws.delete_rows(row_idx)
        except Exception:
            ws.delete_rows(row_idx)
 
def validate_workbooks(wb_src, wb_tpl):
    if SOURCE_SHEET not in wb_src.sheetnames:
        raise ValueError(f"В файле-источнике не найден лист '{SOURCE_SHEET}'.")
 
    for sheet_name in (TEMPLATE_SHEET, STYLE_SOURCE_SHEET):
        if sheet_name not in wb_tpl.sheetnames:
            raise ValueError(f"В файле-шаблоне не найден лист '{sheet_name}'.")
 
def generate(source_path, template_path, operation):
    if not os.path.exists(source_path):
        raise FileNotFoundError("Файл-источник не найден.")
 
    if not os.path.exists(template_path):
        raise FileNotFoundError("Файл-шаблон не найден.")
 
    wb_src = load_workbook(source_path, data_only=True)
    wb_tpl_check = load_workbook(template_path)
    validate_workbooks(wb_src, wb_tpl_check)
    wb_tpl_check.close()
 
    ws_src = wb_src[SOURCE_SHEET]
 
    green_rows = collect_green_rows(ws_src)
    if not green_rows:
        wb_src.close()
        raise ValueError(
            f"Не найдено ни одной строки с цветом новинок {TARGET_GREEN_RGB} в столбце F."
        )
 
    g6_date = next_month_first_day()
    g6_text = g6_date.strftime("%d.%m.%Y")
 
    output_path = build_output_path_from_template(template_path, g6_text)
    create_template_copy(template_path, output_path)
 
    wb_out = load_workbook(output_path)
 
    ws_for_check = wb_out[TEMPLATE_SHEET]
    ws_discounts = wb_out[STYLE_SOURCE_SHEET]
 
    clear_old_rows(ws_for_check)
 
    fill_static_part(ws_for_check, operation, g6_text)
    fill_static_part(ws_discounts, operation, g6_text)

    customer_map = build_customer_map(ws_src)
    nomenclature_map = build_nomenclature_map(ws_src)
 
    start_row = 9
    block_size = len(CODES) * 2  
 
    for src_row in green_rows:
        material = ws_src.cell(src_row, 6).value   
        base_percent = normalize_percent(ws_src.cell(src_row, 13).value)  
 
        y603_map = {}
        for code in CODES:
            y603_map[code] = get_y603_value(ws_src, src_row, code, customer_map)
 
        write_block(
            ws=ws_for_check,
            wb_out=wb_out,
            start_row=start_row,
            material=material,
            base_percent=base_percent,
            y603_map=y603_map,
            nomenclature_map=nomenclature_map
        )
 
        start_row += block_size

    delete_zero_rows(ws_for_check)
    fill_nomenclature_formula(ws_for_check, path=source_path)
    force_data_fill(ws_for_check, start_row=9, end_row=ws_for_check.max_row)
    wb_out.save(output_path)
    wb_out.close()
    wb_src.close()
 
    return output_path, len(green_rows)
 
class App(tk.Tk):
    def __init__(self):
        super().__init__()
 
        self.title("Формирование расценки новинок")
        self.geometry("760x250")
        self.resizable(False, False)
 
        self.src = tk.StringVar()
        self.tpl = tk.StringVar()
        self.op = tk.StringVar(value=OPERATIONS[0])
 
        self.build_ui()
 
    def build_ui(self):
        pad = {"padx": 8, "pady": 8}
 
        ttk.Label(self, text="Файл-источник:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(self, textvariable=self.src, width=62).grid(row=0, column=1, **pad)
        ttk.Button(self, text="Выбрать...", command=self.pick_src).grid(row=0, column=2, **pad)
 
        ttk.Label(self, text="Файл-шаблон:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(self, textvariable=self.tpl, width=62).grid(row=1, column=1, **pad)
        ttk.Button(self, text="Выбрать...", command=self.pick_tpl).grid(row=1, column=2, **pad)
 
        ttk.Label(self, text="Тип операции:").grid(row=2, column=0, sticky="w", **pad)
        cmb = ttk.Combobox(
            self,
            textvariable=self.op,
            values=OPERATIONS,
            state="readonly",
            width=28
        )
        cmb.grid(row=2, column=1, sticky="w", **pad)
 
        ttk.Button(self, text="Сформировать файл", command=self.run).grid(
            row=3, column=1, sticky="w", padx=8, pady=20
        )
 
    def pick_src(self):
        path = filedialog.askopenfilename(
            title="Выберите файл-источник",
            filetypes=[("Excel файлы", "*.xlsx *.xlsm"), ("Все файлы", "*.*")]
        )
        if path:
            self.src.set(path)
 
    def pick_tpl(self):
        path = filedialog.askopenfilename(
            title="Выберите файл-шаблон",
            filetypes=[("Excel файлы", "*.xlsx *.xlsm"), ("Все файлы", "*.*")]
        )
        if path:
            self.tpl.set(path)
 
    def run(self):
        source = self.src.get().strip()
        template = self.tpl.get().strip()
        operation = self.op.get().strip()
 
        if not source:
            messagebox.showwarning("Внимание", "Выберите файл-источник.")
            return
 
        if not template:
            messagebox.showwarning("Внимание", "Выберите файл-шаблон.")
            return
 
        if operation not in OPERATIONS:
            messagebox.showwarning("Внимание", "Выберите корректный тип операции.")
            return
 
        try:
            output_path, processed_rows = generate(
                source_path=source,
                template_path=template,
                operation=operation
            )
 
            messagebox.showinfo(
                "Готово",
                f"Файл успешно сформирован.\n\n"
                f"Обработано строк новинок: {processed_rows}\n"
                f"Сохранено в:\n{output_path}"
            )
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
 
if __name__ == "__main__":
    App().mainloop()